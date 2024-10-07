"""파일별 구분하여 스크립트를 진행합니다"""


import warnings
import fitz
import win32com.client as win32
from openpyxl import load_workbook


from module.extract_information import extract_personal_information
from module.masking_file import _masking_hwp


def processing_pdf(folder_path, pdf_file):
    """pdf파일을 처리후, pdf_infos에 모든 결과를 리스트로 저장하여 return합니다"""
    pdf_infos = []
    try:
        doc = fitz.open(pdf_file)
        for page_num in range(doc.page_count):
            page = doc.load_page(page_num)
            text = page.get_text()
            result, _ = extract_personal_information(folder_path,
                                                     pdf_file, text=text, page_num=page_num)
            pdf_infos.extend(result)
    except Exception as e:  # pylint: disable=W0703
        error_log = str(e)
        pdf_infos.extend(
            extract_personal_information(folder_path, pdf_file, error=error_log))
        print(pdf_file, e)

    return pdf_infos


def processing_hwp(folder_path, hwp_file):
    """hwp 파일을 처리 후, hwp_infos에 모든 결과를 리스트로 저장하여 반환합니다"""
    hwp_infos = []
    hwp = None

    try:
        hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
        hwp.RegisterModule("FilePathCheckDLL", "SecurityModule")
        hwp.Open(hwp_file, arg="versionwarning:False;suspendpassword:True")
        hwp.InitScan()

        while True:
            state, text = hwp.GetText()
            hwp.MovePos(201)
            if state in [0, 1]:
                break
            result, is_success = extract_personal_information(folder_path, hwp_file, text=text,
                                                              page_num=hwp.KeyIndicator()[3])
            hwp_infos.extend(result)

            if is_success and text is not None:
                _masking_hwp(hwp, text)

    except Exception as e:  # pylint: disable=W0703
        error_log = str(e)
        hwp_infos.extend(
            extract_personal_information(folder_path, hwp_file, error=error_log))
        print(hwp_file, e)

    finally:
        if hwp:
            hwp.Save(True)
            hwp.ReleaseScan()
            hwp.Quit()

    return hwp_infos


def processing_excel(folder_path, excel_file):
    """엑셀 파일을 처리 후, excel_infos에 모든 결과를 리스트로 저장하여 반환합니다"""
    excel_infos = []

    try:
        warnings.filterwarnings(action='ignore')
        workbook = load_workbook(excel_file)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row_index, row in enumerate(sheet.iter_rows(), start=1):
                for col_index, cell in enumerate(row, start=1):
                    if cell.value is None:
                        continue
                    xls_index = f"[{row_index + 1}, {col_index}]"
                    text = str(cell.value).strip()
                    result, is_success = extract_personal_information(folder_path, excel_file,
                                                                      text=text, page_num=xls_index)
                    excel_infos.extend(result)

                    if is_success and text is not None:
                        cell.value = '*' * \
                            len(text.replace('\r', '').replace('\n', ''))

        # 수정된 내용을 파일에 저장
        workbook.save(excel_file)

    except Exception as e:  # pylint: disable=W0703
        excel_infos.extend(
            extract_personal_information(folder_path, excel_file, error=str(e)))
        print(excel_file, e)

    return excel_infos
