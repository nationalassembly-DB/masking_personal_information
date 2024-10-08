"""
전달받은 개인정보 리스트를 엑셀에 저장합니다
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


def save_infos_to_excel(infos, excel_file):
    """개인정보를 찾은 리스트를 엑셀 파일로 저장합니다."""
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        headers = ["연번", "위원회", "피감기관", "파일명",
                   "확장자", "페이지번호", "유형", "내용", "파일 이상"]
        header_color = PatternFill(start_color='4f81bd',
                                   end_color='4f81bd', fill_type='solid')

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
            ws.cell(row=1, column=col_idx).fill = header_color

    for i in range(0, len(infos), 5000):
        chunk = infos[i:i + 5000]
        for j, info in enumerate(chunk, start=ws.max_row + i):
            ws.append([j] + list(info))
        wb.save(excel_file)

    wb.save(excel_file)
